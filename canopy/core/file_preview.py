"""
Read-only file preview helpers for text and spreadsheet attachments.

This module intentionally does not execute embedded spreadsheet code or VBA.
It extracts a bounded preview suitable for inline Canopy rendering.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dependency declared in pyproject
    load_workbook = None


SPREADSHEET_MIME_TYPES = {
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
}

SPREADSHEET_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm"}
MARKDOWN_EXTENSIONS = {".md", ".markdown"}
TEXT_PREVIEW_EXTENSIONS = {
    ".md",
    ".markdown",
    ".txt",
    ".log",
    ".json",
    ".py",
    ".js",
    ".ts",
    ".csv",
    ".tsv",
    ".yaml",
    ".yml",
    ".xml",
    ".tex",
    ".html",
    ".css",
    ".sh",
    ".bat",
    ".cfg",
    ".ini",
    ".toml",
}
TEXT_PREVIEW_MIME_PREFIXES = ("text/",)
TEXT_PREVIEW_MIME_TYPES = {
    "application/json",
    "application/xml",
    "application/x-yaml",
    "application/javascript",
    "application/typescript",
    "text/x-tex",
    "application/x-latex",
}

MAX_TEXT_PREVIEW_BYTES = 512 * 1024
MAX_TEXT_PREVIEW_CHARS = 50_000
MAX_SPREADSHEET_PREVIEW_BYTES = 12 * 1024 * 1024
MAX_SHEETS = 3
MAX_ROWS = 60
MAX_COLS = 14
MAX_CELL_CHARS = 160


def _file_extension(filename: str | None) -> str:
    return Path(filename or "").suffix.lower()


def is_markdown_previewable(filename: str | None, content_type: str | None) -> bool:
    ext = _file_extension(filename)
    ctype = str(content_type or "").lower()
    return ext in MARKDOWN_EXTENSIONS or ctype in {"text/markdown", "text/x-markdown"}


def is_spreadsheet_previewable(filename: str | None, content_type: str | None) -> bool:
    ext = _file_extension(filename)
    ctype = str(content_type or "").lower()
    return ext in SPREADSHEET_EXTENSIONS or ctype in SPREADSHEET_MIME_TYPES


def is_text_previewable(filename: str | None, content_type: str | None) -> bool:
    if is_spreadsheet_previewable(filename, content_type):
        return False
    ext = _file_extension(filename)
    ctype = str(content_type or "").lower()
    if ext in TEXT_PREVIEW_EXTENSIONS:
        return True
    if ctype in TEXT_PREVIEW_MIME_TYPES:
        return True
    return any(ctype.startswith(prefix) for prefix in TEXT_PREVIEW_MIME_PREFIXES)


def _decode_text_bytes(file_data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16"):
        try:
            return file_data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_data.decode("utf-8", errors="replace")


def _truncate_text(value: str, limit: int) -> tuple[str, bool]:
    if len(value) <= limit:
        return value, False
    return value[:limit], True


def _serialize_cell(value: Any) -> dict[str, Any]:
    if value is None:
        return {"display": "", "kind": "empty"}
    if isinstance(value, bool):
        return {"display": "TRUE" if value else "FALSE", "kind": "boolean"}
    if isinstance(value, int):
        return {"display": str(value), "kind": "number"}
    if isinstance(value, float):
        if value.is_integer():
            return {"display": str(int(value)), "kind": "number"}
        return {"display": format(value, ".15g"), "kind": "number"}
    if isinstance(value, datetime):
        return {"display": value.isoformat(sep=" ", timespec="seconds"), "kind": "datetime"}
    if isinstance(value, date):
        return {"display": value.isoformat(), "kind": "date"}
    if isinstance(value, time):
        return {"display": value.isoformat(timespec="seconds"), "kind": "time"}

    display = str(value)
    truncated_display, truncated = _truncate_text(display, MAX_CELL_CHARS)
    cell = {"display": truncated_display, "kind": "text"}
    if truncated:
        cell["truncated"] = True
        cell["full_length"] = len(display)
    return cell


def _build_text_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    preview_bytes = file_data[:MAX_TEXT_PREVIEW_BYTES]
    text = _decode_text_bytes(preview_bytes)
    text, text_truncated = _truncate_text(text, MAX_TEXT_PREVIEW_CHARS)
    kind = "markdown" if is_markdown_previewable(filename, content_type) else "text"
    return {
        "previewable": True,
        "kind": kind,
        "text": text,
        "truncated": len(file_data) > MAX_TEXT_PREVIEW_BYTES or text_truncated,
        "limits": {
            "max_bytes": MAX_TEXT_PREVIEW_BYTES,
            "max_chars": MAX_TEXT_PREVIEW_CHARS,
        },
    }


def _csv_rows(text: str, filename: str) -> tuple[list[list[dict[str, Any]]], int, int]:
    sample = text[:8192]
    delimiter = "\t" if _file_extension(filename) == ".tsv" else ","
    try:
        sniffed = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = sniffed.delimiter
    except Exception:
        pass

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[dict[str, Any]]] = []
    total_rows = 0
    total_cols = 0
    for raw_row in reader:
        total_rows += 1
        total_cols = max(total_cols, len(raw_row))
        if len(rows) >= MAX_ROWS:
            continue
        cooked = [_serialize_cell(value) for value in raw_row[:MAX_COLS]]
        while cooked and cooked[-1]["kind"] == "empty":
            cooked.pop()
        rows.append(cooked)
    return rows, total_rows, total_cols


def _build_csv_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    preview_bytes = file_data[:MAX_TEXT_PREVIEW_BYTES]
    text = _decode_text_bytes(preview_bytes)
    rows, total_rows, total_cols = _csv_rows(text, filename)
    return {
        "previewable": True,
        "kind": "spreadsheet",
        "macro_enabled": False,
        "sheets": [
            {
                "name": Path(filename or "Sheet1").stem or "Sheet1",
                "rows": rows,
                "row_count": total_rows,
                "col_count": total_cols,
                "preview_row_count": len(rows),
                "preview_col_count": min(MAX_COLS, max((len(row) for row in rows), default=0)),
                "truncated_rows": total_rows > MAX_ROWS,
                "truncated_cols": total_cols > MAX_COLS,
            }
        ],
        "sheet_count": 1,
        "truncated": len(file_data) > MAX_TEXT_PREVIEW_BYTES or total_rows > MAX_ROWS or total_cols > MAX_COLS,
        "limits": {
            "max_bytes": MAX_TEXT_PREVIEW_BYTES,
            "max_sheets": 1,
            "max_rows": MAX_ROWS,
            "max_cols": MAX_COLS,
        },
    }


def _build_workbook_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    if load_workbook is None:
        return {
            "previewable": False,
            "kind": "spreadsheet",
            "error": "Spreadsheet preview dependency is unavailable on this Canopy instance.",
        }

    if len(file_data) > MAX_SPREADSHEET_PREVIEW_BYTES:
        return {
            "previewable": False,
            "kind": "spreadsheet",
            "error": (
                f"Spreadsheet preview is limited to {MAX_SPREADSHEET_PREVIEW_BYTES // (1024 * 1024)} MB. "
                "Download the file to inspect the full workbook."
            ),
        }

    workbook = load_workbook(
        io.BytesIO(file_data),
        read_only=True,
        data_only=True,
        keep_vba=False,
    )
    sheets: list[dict[str, Any]] = []
    total_sheet_count = len(workbook.worksheets)

    for worksheet in workbook.worksheets[:MAX_SHEETS]:
        rows: list[list[dict[str, Any]]] = []
        preview_col_count = 0
        for raw_row in worksheet.iter_rows(min_row=1, max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True):
            cooked = [_serialize_cell(value) for value in raw_row]
            while cooked and cooked[-1]["kind"] == "empty":
                cooked.pop()
            rows.append(cooked)
            preview_col_count = max(preview_col_count, len(cooked))

        row_count = int(getattr(worksheet, "max_row", 0) or 0)
        col_count = int(getattr(worksheet, "max_column", 0) or 0)
        sheets.append(
            {
                "name": worksheet.title,
                "rows": rows,
                "row_count": row_count,
                "col_count": col_count,
                "preview_row_count": len(rows),
                "preview_col_count": min(MAX_COLS, preview_col_count),
                "truncated_rows": row_count > MAX_ROWS,
                "truncated_cols": col_count > MAX_COLS,
            }
        )

    try:
        workbook.close()
    except Exception:
        pass

    macro_enabled = _file_extension(filename) == ".xlsm" or str(content_type or "").lower() == (
        "application/vnd.ms-excel.sheet.macroenabled.12"
    )
    return {
        "previewable": True,
        "kind": "spreadsheet",
        "macro_enabled": macro_enabled,
        "sheets": sheets,
        "sheet_count": total_sheet_count,
        "truncated": total_sheet_count > MAX_SHEETS or any(sheet["truncated_rows"] or sheet["truncated_cols"] for sheet in sheets),
        "warning": (
            "Workbook preview is read-only. Canopy never executes spreadsheet macros or VBA."
            if macro_enabled else None
        ),
        "limits": {
            "max_bytes": MAX_SPREADSHEET_PREVIEW_BYTES,
            "max_sheets": MAX_SHEETS,
            "max_rows": MAX_ROWS,
            "max_cols": MAX_COLS,
        },
    }


def build_file_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    if is_spreadsheet_previewable(filename, content_type):
        if _file_extension(filename) in {".csv", ".tsv"} or str(content_type or "").lower() == "text/csv":
            return _build_csv_preview(file_data, filename, content_type)
        return _build_workbook_preview(file_data, filename, content_type)
    if is_text_previewable(filename, content_type):
        return _build_text_preview(file_data, filename, content_type)
    return {
        "previewable": False,
        "kind": "unsupported",
        "error": "Inline preview is not available for this file type.",
    }
